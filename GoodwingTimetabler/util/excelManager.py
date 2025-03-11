import pandas as pd

# ExcelScheduleManager
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.filters import FilterColumn, Filters
import datetime as dt
from typing import List, Dict
from collections import defaultdict
import os
import sys

sys.path.append(os.path.abspath(os.path.dirname(os.path.dirname(__file__))))

from GoodwingTimetabler.csp import University

def createCSV(gen_dir: str = './GoodwingTimetabler/UniversityInstance/'):
    sheets = ['University', 'Timeslots', 'Promotions', 'Subjects', 'Teachers', 'Rooms']
    dfs = []
    for idx, sheet_name in enumerate(sheets):
        dfs.append(pd.read_excel(gen_dir+'University.xlsx', sheet_name))
        dfs[idx].to_csv(f'{gen_dir}csv/{sheet_name}.csv', index=False)

    # Print the CSV
    #for df in dfs:
    #    pd.set_option('display.max_columns', 1000)
    #    pd.set_option('display.width', 1000)
    #    print(df)
    #    print('\n')

class ExcelScheduleManager:
    def __init__(self, university: University, generated_courses):
        self.university = university
        self.courses = generated_courses
        self.wb = Workbook()
        self.time_slots = university.time_ranges
        
    def format_time(self, time_obj):
        return time_obj.strftime("%H:%M")

    def get_week_number(self, date):
        return (date - self.university.timeslots[0].day).days // 7 + 1

    def setup_sheet_structure(self, ws):
        # Headers
        headers = ['Week', 'Day', 'Time Slot', 'Subject', 'Teacher', 'Room']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add filters
        ws.auto_filter.ref = f'A1:F1'
        
        # Set column widths
        dim_holder = DimensionHolder(worksheet=ws)
        
        widths = [10, 15, 20, 30, 25, 15]  # Adjusted widths for each column
        for col, width in enumerate(widths, 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)
        
        ws.column_dimensions = dim_holder

    def add_course_to_sheet(self, ws, row, course, week_num):
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_of_week = day_names[course.timeslot.day.weekday()]
        
        # Data to insert
        data = [
            week_num,
            day_of_week,
            f"{self.format_time(course.timeslot.start)} - {self.format_time(course.timeslot.end)}",
            course.subject.name,
            f"{course.teacher.first_name} {course.teacher.last_name}",
            course.room.name
        ]
        
        # Insert data
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add color to subject cell
            if col == 4:  # Subject column
                cell.fill = PatternFill(start_color=course.subject.color, 
                                      end_color=course.subject.color, 
                                      fill_type="solid")
                # Adjust font color for better readability
                brightness = sum(int(course.subject.color[i:i+2], 16) for i in (0, 2, 4)) / 3
                cell.font = Font(color="FFFFFF" if brightness < 128 else "000000")

        # Add border
        border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border

    def create_group_schedule(self, group_name, group_courses):
        ws = self.wb.create_sheet(title=f"Group_{group_name}")
        self.setup_sheet_structure(ws)
        
        row = 2
        for course in sorted(group_courses, 
                           key=lambda x: (self.get_week_number(x.timeslot.day), 
                                        x.timeslot.day, 
                                        x.timeslot.start)):
            week_num = self.get_week_number(course.timeslot.day)
            self.add_course_to_sheet(ws, row, course, week_num)
            row += 1

    def create_teacher_schedule(self, teacher, teacher_courses):
        ws = self.wb.create_sheet(title=f"Teacher_{teacher.last_name}")
        self.setup_sheet_structure(ws)
        
        row = 2
        for course in sorted(teacher_courses, 
                           key=lambda x: (self.get_week_number(x.timeslot.day), 
                                        x.timeslot.day, 
                                        x.timeslot.start)):
            week_num = self.get_week_number(course.timeslot.day)
            self.add_course_to_sheet(ws, row, course, week_num)
            row += 1

    def create_room_schedule(self, room, room_courses):
        ws = self.wb.create_sheet(title=f"Room_{room.name}")
        self.setup_sheet_structure(ws)
        
        row = 2
        for course in sorted(room_courses, 
                           key=lambda x: (self.get_week_number(x.timeslot.day), 
                                        x.timeslot.day, 
                                        x.timeslot.start)):
            week_num = self.get_week_number(course.timeslot.day)
            self.add_course_to_sheet(ws, row, course, week_num)
            row += 1

    def create_statistics_sheet(self):
        ws = self.wb.create_sheet(title="Statistics")
        
        # Style for headers
        header_style = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Group Statistics
        ws['A1'] = "Group Statistics"
        ws['A1'].font = header_style
        
        headers = ['Group', 'Total Hours', 'Subjects', 'Teachers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
        
        row = 3
        group_stats = defaultdict(lambda: {'hours': 0, 'subjects': set(), 'teachers': set()})
        
        for course in self.courses:
            stats = group_stats[course.group.name]
            duration = (course.timeslot.end.hour - course.timeslot.start.hour + 
                       (course.timeslot.end.minute - course.timeslot.start.minute) / 60)
            stats['hours'] += duration
            stats['subjects'].add(course.subject.name)
            stats['teachers'].add(f"{course.teacher.first_name} {course.teacher.last_name}")
        
        for group, stats in group_stats.items():
            ws.cell(row=row, column=1, value=group)
            ws.cell(row=row, column=2, value=f"{stats['hours']:.1f}")
            ws.cell(row=row, column=3, value=len(stats['subjects']))
            ws.cell(row=row, column=4, value=len(stats['teachers']))
            row += 1
        
        # Teacher Statistics
        row += 2
        ws.cell(row=row, column=1, value="Teacher Statistics").font = header_style
        row += 1
        
        headers = ['Teacher', 'Total Hours', 'Groups', 'Subjects']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
        
        row += 1
        teacher_stats = defaultdict(lambda: {'hours': 0, 'groups': set(), 'subjects': set()})
        
        for course in self.courses:
            teacher = f"{course.teacher.first_name} {course.teacher.last_name}"
            stats = teacher_stats[teacher]
            duration = (course.timeslot.end.hour - course.timeslot.start.hour + 
                       (course.timeslot.end.minute - course.timeslot.start.minute) / 60)
            stats['hours'] += duration
            stats['groups'].add(course.group.name)
            stats['subjects'].add(course.subject.name)
        
        for teacher, stats in teacher_stats.items():
            ws.cell(row=row, column=1, value=teacher)
            ws.cell(row=row, column=2, value=f"{stats['hours']:.1f}")
            ws.cell(row=row, column=3, value=len(stats['groups']))
            ws.cell(row=row, column=4, value=len(stats['subjects']))
            row += 1
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def generate_excel_schedule(self, output_path):
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Group schedules
        group_courses = defaultdict(list)
        for course in self.courses:
            group_courses[course.group.name].append(course)
        
        for group_name, courses in group_courses.items():
            self.create_group_schedule(group_name, courses)
        
        # Teacher schedules
        teacher_courses = defaultdict(list)
        for course in self.courses:
            teacher_courses[course.teacher].append(course)
        
        for teacher, courses in teacher_courses.items():
            self.create_teacher_schedule(teacher, courses)
        
        # Room schedules
        room_courses = defaultdict(list)
        for course in self.courses:
            room_courses[course.room].append(course)
        
        for room, courses in room_courses.items():
            self.create_room_schedule(room, courses)
        
        # Statistics sheet
        self.create_statistics_sheet()
        
        # Save the workbook
        self.wb.save(output_path)
